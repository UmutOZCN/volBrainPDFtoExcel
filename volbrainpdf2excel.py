# ---------------------------------------------------------
# volBrain PDF Excel Converter (TR/EN)
# Developer: Umut Özcan
# Date: 2025
# Description: PDF reports -> Excel converter (Bilingual UI)
# ---------------------------------------------------------

DEVELOPER_SIGNATURE = "Developed by Umut Özcan - 2025"

import os
import re
from datetime import datetime
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# ---------------- Localization ----------------
LANG = {
    'tr': {
        'window_title': "volBrain PDF to Excel Dönüştürücü",
        'app_title': "volBrain PDF'den Excel'e Veri Dönüştürücü",
        'pdf_folder': "PDF Klasörü:",
        'browse': "Gözat",
        'excel_file': "Excel Çıktı Dosyası:",
        'status_ready': "Hazır",
        'convert': "Dönüştür",
        'menu_help': "Yardım",
        'menu_about': "Hakkında",
        'about_text': "volBrain PDF Excel Dönüştürücü\n\nGeliştirici: Umut Özcan\n© 2025",
        'err_select_pdf': "Lütfen bir PDF klasörü seçin!",
        'err_select_xlsx': "Lütfen bir Excel çıktı dosyası seçin!",
        'err_pdf_missing': "Seçilen PDF klasörü mevcut değil!",
        'err_no_pdf': "PDF dosyası bulunamadı!",
        'processing': "İşleniyor",
        'creating_excel': "Excel dosyası oluşturuluyor...",
        'ok_title': "Başarılı",
        'ok_body': "Dönüştürme tamamlandı!\n\nBaşarılı: {ok}\nBaşarısız: {fail}\n\nExcel dosyası: {path}",
        'err_excel': "Excel dosyası oluşturulamadı!",
        'err_conv': "Dönüştürme sırasında hata oluştu: {err}",
        'status_done': "Tamamlandı - {count} dosya işlendi",
        'filedlg_pdf_title': "PDF Dosyalarının Bulunduğu Klasörü Seçin",
        'filedlg_xlsx_title': "Excel Dosyasını Kaydet",
        'sheet_title': "HASTA VERİLERİ",
        'created_at': "Oluşturulma Tarihi: {dt}",
        'total_records': "Toplam Kayıt Sayısı: {n}",
        'language_label': "Dil / Language:",
        'headers': {
            "Hasta ID": "Hasta ID",
            "Ad Soyad": "Ad Soyad",
            "Yaş": "Yaş",
            "Cinsiyet": "Cinsiyet",
            "White Matter (WM)": "White Matter (WM)",
            "Grey Matter (GM)": "Grey Matter (GM)",
            "Cerebro Spinal Fluid (CSF)": "Cerebro Spinal Fluid (CSF)",
            "Brain (WM + GM)": "Brain (WM + GM)",
            "Intracranial Cavity (IC)": "Intracranial Cavity (IC)",
            "Cerebrum Total": "Cerebrum Total",
            "Cerebrum Right": "Cerebrum Right",
            "Cerebrum Left": "Cerebrum Left",
            "Cerebrum Asym": "Cerebrum Asym",
            "Cerebellum Total": "Cerebellum Total",
            "Cerebellum Right": "Cerebellum Right",
            "Cerebellum Left": "Cerebellum Left",
            "Cerebellum Asym": "Cerebellum Asym",
            "Brainstem Total": "Brainstem Total",
            "Lateral ventricles Total": "Lateral ventricles Total",
            "Caudate Total": "Caudate Total",
            "Putamen Total": "Putamen Total",
            "Thalamus Total": "Thalamus Total",
            "Globus Pallidus Total": "Globus Pallidus Total",
            "Hippocampus Total": "Hippocampus Total",
            "Amygdala Total": "Amygdala Total",
            "Accumbens Total": "Accumbens Total"
        }
    },
    'en': {
        'window_title': "volBrain PDF to Excel Converter",
        'app_title': "volBrain PDF to Excel Data Converter",
        'pdf_folder': "PDF Folder:",
        'browse': "Browse",
        'excel_file': "Excel Output File:",
        'status_ready': "Ready",
        'convert': "Convert",
        'menu_help': "Help",
        'menu_about': "About",
        'about_text': "volBrain PDF to Excel Converter\n\nDeveloper: Umut Özcan\n© 2025",
        'err_select_pdf': "Please select a PDF folder!",
        'err_select_xlsx': "Please select an Excel output file!",
        'err_pdf_missing': "Selected PDF folder does not exist!",
        'err_no_pdf': "No PDF files found!",
        'processing': "Processing",
        'creating_excel': "Creating Excel file...",
        'ok_title': "Success",
        'ok_body': "Conversion completed!\n\nSuccess: {ok}\nFailed: {fail}\n\nExcel file: {path}",
        'err_excel': "Failed to create Excel file!",
        'err_conv': "An error occurred during conversion: {err}",
        'status_done': "Completed - {count} files processed",
        'filedlg_pdf_title': "Select the Folder Containing PDF Files",
        'filedlg_xlsx_title': "Save Excel File",
        'sheet_title': "PATIENT DATA",
        'created_at': "Created At: {dt}",
        'total_records': "Total Records: {n}",
        'language_label': "Language / Dil:",
        'headers': {
            "Hasta ID": "Patient ID",
            "Ad Soyad": "Full Name",
            "Yaş": "Age",
            "Cinsiyet": "Sex",
            "White Matter (WM)": "White Matter (WM)",
            "Grey Matter (GM)": "Grey Matter (GM)",
            "Cerebro Spinal Fluid (CSF)": "Cerebro Spinal Fluid (CSF)",
            "Brain (WM + GM)": "Brain (WM + GM)",
            "Intracranial Cavity (IC)": "Intracranial Cavity (IC)",
            "Cerebrum Total": "Cerebrum Total",
            "Cerebrum Right": "Cerebrum Right",
            "Cerebrum Left": "Cerebrum Left",
            "Cerebrum Asym": "Cerebrum Asym",
            "Cerebellum Total": "Cerebellum Total",
            "Cerebellum Right": "Cerebellum Right",
            "Cerebellum Left": "Cerebellum Left",
            "Cerebellum Asym": "Cerebellum Asym",
            "Brainstem Total": "Brainstem Total",
            "Lateral ventricles Total": "Lateral ventricles Total",
            "Caudate Total": "Caudate Total",
            "Putamen Total": "Putamen Total",
            "Thalamus Total": "Thalamus Total",
            "Globus Pallidus Total": "Globus Pallidus Total",
            "Hippocampus Total": "Hippocampus Total",
            "Amygdala Total": "Amygdala Total",
            "Accumbens Total": "Accumbens Total"
        }
    }
}

# --------------- Parsing helpers ----------------
def extract_patient_name(pdf_path):
    try:
        filename = os.path.basename(pdf_path)
        filename_no_ext = re.sub(r'\.pdf$', '', filename, flags=re.I)
        pattern = r'[^a-zA-ZçÇğĞıİöÖşŞüÜ\s\-]'
        name = re.sub(pattern, ' ', filename_no_ext)
        name = re.sub(r'\s+', ' ', name).strip()
        return name or "İsim Çıkarılamadı"
    except Exception:
        return "İsim Çıkarılamadı"

def extract_data_from_pdf(pdf_path):
    data = {
        "Hasta ID": "",
        "Ad Soyad": extract_patient_name(pdf_path),
        "Yaş": "",
        "Cinsiyet": "",
        "White Matter (WM)": "",
        "Grey Matter (GM)": "",
        "Cerebro Spinal Fluid (CSF)": "",
        "Brain (WM + GM)": "",
        "Intracranial Cavity (IC)": "",
        "Cerebrum Total": "",
        "Cerebrum Right": "",
        "Cerebrum Left": "",
        "Cerebrum Asym": "",
        "Cerebellum Total": "",
        "Cerebellum Right": "",
        "Cerebellum Left": "",
        "Cerebellum Asym": "",
        "Brainstem Total": "",
        "Lateral ventricles Total": "",
        "Caudate Total": "",
        "Putamen Total": "",
        "Thalamus Total": "",
        "Globus Pallidus Total": "",
        "Hippocampus Total": "",
        "Amygdala Total": "",
        "Accumbens Total": ""
    }

    try:
        with pdfplumber.open(pdf_path) as pdf:
            all_text = ""
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    all_text += t + "\n"

            all_text = re.sub(r"\u00a0", " ", all_text)
            all_text = re.sub(r"[\t ]+", " ", all_text)

            # Patient ID
            m = re.search(r'Patient ID\s*[:\-]?\s*([^\n]+)', all_text, re.I)
            if m:
                data["Hasta ID"] = re.sub(r'\D', '', m.group(1))
            else:
                m2 = re.search(r'\b(?:Job|ID)\b\s*[:\-]?\s*([^\n]+)', all_text, re.I)
                if m2:
                    data["Hasta ID"] = re.sub(r'\D', '', m2.group(1))
                else:
                    m3 = re.search(r'job\s*(\d+)|job(\d+)', all_text, re.I)
                    if m3:
                        data["Hasta ID"] = (m3.group(1) or m3.group(2))

            # Age
            fn = os.path.basename(pdf_path)
            m_age = re.search(r'(\d{1,3})\s*$', re.sub(r'\.pdf$', '', fn, flags=re.I))
            if m_age:
                data["Yaş"] = m_age.group(1)
            else:
                m_age2 = re.search(r'Age[^\d]{0,10}(\d{1,3})', all_text, re.I | re.S)
                if m_age2:
                    data["Yaş"] = m_age2.group(1).strip()

            # Sex
            m_sex = re.search(r'\bSex\b\s*[:\-]?\s*(Male|Female|Erkek|Kadın|E|K|M|F)', all_text, re.I)
            if m_sex:
                s = m_sex.group(1).strip().lower()
                if s in ("male", "m", "erkek", "e"):
                    data["Cinsiyet"] = "Erkek"
                elif s in ("female", "f", "kadın", "k"):
                    data["Cinsiyet"] = "Kadın"
            else:
                if re.search(r'\bMale\b', all_text, re.I): data["Cinsiyet"] = "Erkek"
                elif re.search(r'\bFemale\b', all_text, re.I): data["Cinsiyet"] = "Kadın"

            # Helpers
            def first_non_percent_decimal(line_text):
                cands = list(re.finditer(r'(\d+(?:[\.,]\d+)?)', line_text))
                for m in cands:
                    tail = line_text[m.end():m.end()+5]
                    if not re.match(r'\s*%', ' ' + tail):
                        return m.group(0)
                return cands[-1].group(0) if cands else None

            def value_after_label(block_text, label_regex):
                m = re.search(label_regex + r'[\s\S]{0,120}', block_text, re.I)
                if not m: return None
                return first_non_percent_decimal(m.group(0))

            # Volumes (normalize to comma-decimal)
            def norm(x): return x.replace(',', '.').replace('.', ',')

            m = re.search(r'White\s*Matter\s*\(WM\).*?(\d+[\.,]?\d*)', all_text, re.I)
            if m: data["White Matter (WM)"] = norm(m.group(1))
            m = re.search(r'Grey\s*Matter\s*\(GM\).*?(\d+[\.,]?\d*)', all_text, re.I)
            if m: data["Grey Matter (GM)"] = norm(m.group(1))
            m = re.search(r'Cerebro\s*Spinal\s*Fluid\s*\(CSF\).*?(\d+[\.,]?\d*)', all_text, re.I)
            if m: data["Cerebro Spinal Fluid (CSF)"] = norm(m.group(1))
            m = re.search(r'Brain\s*\((?:WM\s*\+\s*GM|GM\s*\+\s*WM)\).*?(\d+[\.,]?\d*)', all_text, re.I)
            if m: data["Brain (WM + GM)"] = norm(m.group(1))

            m = re.search(r'Intracranial[\sA-Za-z]*\((?:IC|ICV)\).*?(\d+[\.,]\d+)', all_text, re.I)
            if m: data["Intracranial Cavity (IC)"] = norm(m.group(1))
            else:
                ic_line = re.search(r'^.*Intracranial.*$', all_text, re.I | re.M)
                if ic_line:
                    n = re.search(r'(\d+[\.,]\d+)', ic_line.group(0))
                    if n: data["Intracranial Cavity (IC)"] = norm(n.group(1))

            # Cerebrum
            header_idx = re.search(r'(?mi)^\s*Cerebrum\s+Total.*$', all_text)
            table_line = None
            if header_idx:
                after = all_text[header_idx.end():]
                table_line = re.search(r'(?m)^\s*([-\d\.]+)\([^\)]*\)\s+([-\d\.]+)\([^\)]*\)\s+([-\d\.]+)\([^\)]*\)\s+([-\d\.]+)', after)
            if table_line:
                t, r, l, a = table_line.group(1,2,3,4)
                data["Cerebrum Total"] = t.replace('.', ',')
                data["Cerebrum Right"] = r.replace('.', ',')
                data["Cerebrum Left"]  = l.replace('.', ',')
                data["Cerebrum Asym"]  = a.replace('.', ',')
            else:
                m = re.search(
                    r'Cerebrum\s+Total.*?(?P<total>\d+[\.,]\d+).*?Right.*?(?P<right>\d+[\.,]\d+).*?Left.*?(?P<left>\d+[\.,]\d+).*?Asym.*?(?P<asym>-?\d+[\.,]\d+)',
                    all_text, re.I | re.S
                )
                if m:
                    data["Cerebrum Total"] = norm(m.group('total'))
                    data["Cerebrum Right"] = norm(m.group('right'))
                    data["Cerebrum Left"]  = norm(m.group('left'))
                    data["Cerebrum Asym"]  = norm(m.group('asym'))
                else:
                    block = re.search(r'Cerebrum[\s\S]{0,300}', all_text, re.I)
                    if block:
                        bt = block.group(0)
                        def x(rx):
                            v = value_after_label(bt, rx)
                            return norm(v) if v else ""
                        t = x(r'Cerebrum\s+Total|\bTotal\b')
                        r = x(r'(?:Cerebrum\s+)?Right')
                        l = x(r'(?:Cerebrum\s+)?Left')
                        a = x(r'(?:Cerebrum\s+)?Asym')
                        if t: data["Cerebrum Total"] = t
                        if r: data["Cerebrum Right"] = r
                        if l: data["Cerebrum Left"]  = l
                        if a: data["Cerebrum Asym"]  = a

            # Cerebellum
            cb_header = re.search(r'(?mi)^\s*Cereb(?:e|)l+um\s+Total.*$', all_text)
            table_line = None
            if cb_header:
                after = all_text[cb_header.end():]
                table_line = re.search(r'(?m)^\s*([-\d\.]+)\([^\)]*\)\s+([-\d\.]+)\([^\)]*\)\s+([-\d\.]+)\([^\)]*\)\s+([-\d\.]+)', after)
            if table_line:
                t, r, l, a = table_line.group(1,2,3,4)
                data["Cerebellum Total"] = t.replace('.', ',')
                data["Cerebellum Right"] = r.replace('.', ',')
                data["Cerebellum Left"]  = l.replace('.', ',')
                data["Cerebellum Asym"]  = a.replace('.', ',')
            else:
                m = re.search(
                    r'Cereb(?:e|)l+um\s+Total.*?(?P<total>\d+[\.,]\d+).*?Right.*?(?P<right>\d+[\.,]\d+).*?Left.*?(?P<left>\d+[\.,]\d+).*?Asym.*?(?P<asym>-?\d+[\.,]\d+)',
                    all_text, re.I | re.S
                )
                if m:
                    data["Cerebellum Total"] = norm(m.group('total'))
                    data["Cerebellum Right"] = norm(m.group('right'))
                    data["Cerebellum Left"]  = norm(m.group('left'))
                    data["Cerebellum Asym"]  = norm(m.group('asym'))

            # Brainstem & ventricles & subcorticals
            m = re.search(r'Brainstem\s+Total.*?(\d+[\.,]\d+)', all_text, re.I | re.S)
            if m: data["Brainstem Total"] = norm(m.group(1))

            m = re.search(r'ventricles\s+(\d+[\.,]?\d*)\s*\(', all_text, re.I)
            if m: data["Lateral ventricles Total"] = norm(m.group(1))
            elif re.search(r'ventricles\s+0\.00', all_text, re.I):
                data["Lateral ventricles Total"] = "0,00"

            m = re.search(r'Pallidus\s+(\d+[\.,]?\d*)\s*\(', all_text, re.I)
            if m: data["Globus Pallidus Total"] = norm(m.group(1))

            for s in ["Caudate", "Putamen", "Thalamus", "Hippocampus", "Amygdala", "Accumbens"]:
                m = re.search(s + r"\s+(\d+[\.,]?\d*)\s*\(", all_text, re.I)
                if m: data[f"{s} Total"] = norm(m.group(1))

    except Exception as e:
        print(f"{pdf_path} işlenirken hata oluştu: {e}")

    return data

# --------------- Excel writer ----------------
def save_to_excel(data_list, output_file, lang_code='tr'):
    if not data_list:
        print("Kaydedilecek veri yok")
        return False

    L = LANG[lang_code]
    headers_map = L['headers']

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Hasta Verileri" if lang_code == 'tr' else "Patient Data"

        ws['A1'] = L['sheet_title']; ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:Z1'); ws.row_dimensions[1].height = 28

        ts = datetime.now().strftime('%d/%m/%Y %H:%M') if lang_code=='tr' else datetime.now().strftime('%Y-%m-%d %H:%M')
        ws['A2'] = L['created_at'].format(dt=ts)
        ws['A3'] = L['total_records'].format(n=len(data_list))

        original_keys = list(data_list[0].keys())
        display_headers = [headers_map.get(k, k) for k in original_keys]

        for i, h in enumerate(display_headers, 1):
            c = ws.cell(row=5, column=i, value=h)
            c.font = Font(bold=True)
            c.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        for r, row in enumerate(data_list, 6):
            for c, key in enumerate(original_keys, 1):
                ws.cell(row=r, column=c, value=row.get(key, ""))

        for col_idx in range(1, len(display_headers)+1):
            maxlen = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_row=5, max_row=len(data_list)+5, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        v = "" if cell.value is None else str(cell.value)
                        maxlen = max(maxlen, len(v))
                    except Exception:
                        pass
            ws.column_dimensions[col_letter].width = max(16, min(maxlen+2, 32))

        last_col = get_column_letter(len(display_headers))
        ws.auto_filter.ref = f"A5:{last_col}{len(data_list)+5}"

        wb.save(output_file)
        return True
    except Exception as e:
        print(f"Excel kaydetme hatası: {e}")
        return False

# --------------- UI ----------------
class PDFtoExcelConverter:
    def __init__(self, root):
        self.root = root
        self.lang_code = tk.StringVar(value='en')  # default English
        self._apply_language()

        self.style = ttk.Style()
        self.style.configure("TButton", padding=6, relief="flat")
        self.style.configure("Title.TLabel", font=("Arial", 16, "bold"))

        topbar = ttk.Frame(root); topbar.pack(pady=(10, 0), padx=20, fill="x")
        ttk.Label(topbar, text=LANG[self.lang_code.get()]['language_label']).pack(side="left")
        self.lang_combo = ttk.Combobox(topbar, values=["Türkçe", "English"], state="readonly", width=12)
        self.lang_combo.current(1)  # English
        self.lang_combo.pack(side="left", padx=(6, 0))
        self.lang_combo.bind("<<ComboboxSelected>>", self._on_language_change)

        self.title_label = ttk.Label(root, text=LANG[self.lang_code.get()]['app_title'], style="Title.TLabel")
        self.title_label.pack(pady=14)

        self.pdf_frame = ttk.Frame(root); self.pdf_frame.pack(pady=6, padx=20, fill="x")
        self.pdf_label = ttk.Label(self.pdf_frame, text=LANG[self.lang_code.get()]['pdf_folder']); self.pdf_label.pack(anchor="w")
        self.pdf_btn_frame = ttk.Frame(self.pdf_frame); self.pdf_btn_frame.pack(fill="x", pady=5)
        self.pdf_path = tk.StringVar()
        self.pdf_entry = ttk.Entry(self.pdf_btn_frame, textvariable=self.pdf_path); self.pdf_entry.pack(side="left", fill="x", expand=True, padx=(0,5))
        self.pdf_btn = ttk.Button(self.pdf_btn_frame, text=LANG[self.lang_code.get()]['browse'], command=self.select_pdf_folder); self.pdf_btn.pack(side="right")

        self.excel_frame = ttk.Frame(root); self.excel_frame.pack(pady=6, padx=20, fill="x")
        self.excel_label = ttk.Label(self.excel_frame, text=LANG[self.lang_code.get()]['excel_file']); self.excel_label.pack(anchor="w")
        self.excel_btn_frame = ttk.Frame(self.excel_frame); self.excel_btn_frame.pack(fill="x", pady=5)
        self.excel_path = tk.StringVar()
        self.excel_entry = ttk.Entry(self.excel_btn_frame, textvariable=self.excel_path); self.excel_entry.pack(side="left", fill="x", expand=True, padx=(0,5))
        self.excel_btn = ttk.Button(self.excel_btn_frame, text=LANG[self.lang_code.get()]['browse'], command=self.select_excel_file); self.excel_btn.pack(side="right")

        self.progress = ttk.Progressbar(root, orient="horizontal", length=500, mode="determinate"); self.progress.pack(pady=16)
        self.status_label = ttk.Label(root, text=LANG[self.lang_code.get()]['status_ready']); self.status_label.pack(pady=4)

        self.convert_btn = ttk.Button(root, text=LANG[self.lang_code.get()]['convert'], command=self.convert); self.convert_btn.pack(pady=14)

        self._set_default_paths()

    def _apply_language(self):
        L = LANG[self.lang_code.get()]
        self.root.title(L['window_title'])

    def _refresh_texts(self):
        L = LANG[self.lang_code.get()]
        self.root.title(L['window_title'])
        self.title_label.config(text=L['app_title'])
        self.pdf_label.config(text=L['pdf_folder'])
        self.pdf_btn.config(text=L['browse'])
        self.excel_label.config(text=L['excel_file'])
        self.excel_btn.config(text=L['browse'])
        self.status_label.config(text=L['status_ready'])
        self.convert_btn.config(text=L['convert'])

        try:
            self._rebuild_menu()
        except Exception:
            pass

    def _rebuild_menu(self):
        """Re-create menubar with current language labels."""
        if not hasattr(self, "menubar"):
            return
        L = LANG[self.lang_code.get()]

        try:
            self.menubar.delete(0, "end")
        except Exception:
            pass

        self.helpmenu = tk.Menu(self.menubar, tearoff=0)
        self.helpmenu.add_command(
            label=L['menu_about'],
            command=show_about  
        )
        self.menubar.add_cascade(label=L['menu_help'], menu=self.helpmenu)
        self.root.config(menu=self.menubar)

    def _on_language_change(self, _evt=None):
        chosen = self.lang_combo.get()
        self.lang_code.set('tr' if chosen == "Türkçe" else 'en')
        self._refresh_texts()

    def _resolve_desktop_path(self):
        home = os.path.expanduser("~")
        candidates = [
            os.path.join(home, "Desktop"),
            os.path.join(home, "Masaüstü"),
            os.path.join(home, "OneDrive", "Desktop"),
            os.path.join(home, "OneDrive", "Masaüstü"),
        ]
        for p in candidates:
            if os.path.isdir(p): return p
        return candidates[0]

    def _set_default_paths(self):
        d = self._resolve_desktop_path()
        self.pdf_path.set(os.path.join(d, "pdf_files"))
        self.excel_path.set(os.path.join(d, "beyin_hacim_verileri.xlsx"))

    def select_pdf_folder(self):
        L = LANG[self.lang_code.get()]
        p = filedialog.askdirectory(title=L['filedlg_pdf_title'])
        if p: self.pdf_path.set(p)

    def select_excel_file(self):
        L = LANG[self.lang_code.get()]
        p = filedialog.asksaveasfilename(
            title=L['filedlg_xlsx_title'],
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if p: self.excel_path.set(p)

    def convert(self):
        L = LANG[self.lang_code.get()]
        pdf_folder = self.pdf_path.get()
        excel_file = self.excel_path.get()

        if not pdf_folder:
            messagebox.showerror(L['menu_help'], L['err_select_pdf']); return
        if not excel_file:
            messagebox.showerror(L['menu_help'], L['err_select_xlsx']); return
        if not os.path.exists(pdf_folder):
            messagebox.showerror(L['menu_help'], L['err_pdf_missing']); return

        os.makedirs(os.path.dirname(excel_file), exist_ok=True)

        try:
            self.convert_btn.config(state="disabled")
            self.status_label.config(text=f"{L['processing']}: ...")
            self.root.update()

            pdf_files = [f for f in os.listdir(pdf_folder) if f.lower().endswith('.pdf')]
            if not pdf_files:
                messagebox.showerror(L['menu_help'], L['err_no_pdf']); return

            self.progress["maximum"] = len(pdf_files); self.progress["value"] = 0

            data_list, ok, fail = [], 0, 0
            for i, fn in enumerate(pdf_files):
                self.status_label.config(text=f"{L['processing']}: {fn} ({i+1}/{len(pdf_files)})")
                self.root.update()

                try:
                    d = extract_data_from_pdf(os.path.join(pdf_folder, fn))
                    if d: data_list.append(d); ok += 1
                    else: fail += 1
                except Exception as e:
                    print(f"{fn} işlenirken hata: {e}"); fail += 1

                self.progress["value"] = i+1; self.root.update()

            if data_list:
                self.status_label.config(text=L['creating_excel']); self.root.update()
                if save_to_excel(data_list, excel_file, self.lang_code.get()):
                    messagebox.showinfo(L['ok_title'], L['ok_body'].format(ok=ok, fail=fail, path=excel_file))
                    self.status_label.config(text=L['status_done'].format(count=ok))
                else:
                    messagebox.showerror(L['menu_help'], L['err_excel'])
            else:
                messagebox.showerror(L['menu_help'], L['err_no_pdf'])

        except Exception as e:
            messagebox.showerror(L['menu_help'], L['err_conv'].format(err=str(e)))
        finally:
            self.convert_btn.config(state="normal")

# --------------- About dialog ----------------
def show_about():
    code = getattr(app, 'lang_code', tk.StringVar(value='en')).get()
    L = LANG.get(code, LANG['en'])
    messagebox.showinfo(L['menu_about'], L['about_text'])

# --------------- Main ----------------
if __name__ == "__main__":
    root = tk.Tk()

    menubar = tk.Menu(root)
    helpmenu = tk.Menu(menubar, tearoff=0)
    helpmenu.add_command(label=LANG['tr']['menu_about'], command=show_about)
    menubar.add_cascade(label=LANG['tr']['menu_help'], menu=helpmenu)
    root.config(menu=menubar)

    app = PDFtoExcelConverter(root)
    app.menubar = menubar
    app.helpmenu = helpmenu
    app._rebuild_menu()  

    root.mainloop()
